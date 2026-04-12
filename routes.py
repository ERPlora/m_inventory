"""
Inventory module HTMX views -- FastAPI router.

Replaces Django views.py + urls.py. Uses @htmx_view decorator.
Mounted at /m/inventory/ by ModuleRuntime.
"""

from __future__ import annotations

import csv as csv_mod
import io
import uuid
from decimal import Decimal

from fastapi import APIRouter, Request
from fastapi.responses import JSONResponse, Response, StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import selectinload

from app.core.db.query import HubQuery
from app.core.db.transactions import atomic
from app.core.dependencies import CurrentUser, DbSession, HubId
from app.core.htmx import htmx_redirect, htmx_view

from .models import Category, InventoryConfig, Product

router = APIRouter()


def _q(model, db, hub_id):
    return HubQuery(model, db, hub_id)


async def _get_config(db, hub_id) -> InventoryConfig:
    """Get or create inventory config singleton."""
    q = _q(InventoryConfig, db, hub_id)
    config = await q.first()
    if config is None:
        config = InventoryConfig(hub_id=hub_id)
        db.add(config)
        await db.flush()
    return config


# ============================================================================
# Dashboard
# ============================================================================

@router.get("/")
@htmx_view(module_id="inventory", view_id="dashboard", partial_template="inventory/partials/dashboard_content.html")
async def dashboard(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Dashboard with stats and summary."""
    q = _q(Product, db, hub_id)
    active_q = q.filter(Product.is_active == True)  # noqa: E712

    total_products = await active_q.count()
    products_in_stock = await active_q.filter(Product.stock > 0).count()
    products_low_stock = await active_q.filter(
        Product.stock <= Product.low_stock_threshold
    ).count()
    total_inventory_value = await active_q.sum(Product.stock * Product.price) or 0

    # Top 5 categories
    categories = await _q(Category, db, hub_id).filter(
        Category.is_active == True  # noqa: E712
    ).options(
        selectinload(Category.products),
    ).order_by(Category.order, Category.name).limit(5).all()

    # Low stock products (top 10)
    low_stock_products = await active_q.filter(
        Product.stock <= Product.low_stock_threshold
    ).order_by(Product.stock).limit(10).all()

    currency = "EUR"

    return {
        "total_products": total_products,
        "products_in_stock": products_in_stock,
        "products_low_stock": products_low_stock,
        "total_inventory_value": total_inventory_value,
        "currency": currency,
        "categories": categories,
        "low_stock_products": low_stock_products,
    }


# ============================================================================
# Products List
# ============================================================================

@router.get("/products")
@htmx_view(module_id="inventory", view_id="products", permissions="inventory.view_product")
async def products_list(
    request: Request, db: DbSession, user: CurrentUser, hub_id: HubId,
    search: str = "", order_by: str = "-id", page: int = 1, per_page: int = 25,
):
    """Product list with infinite scroll, search, and sort."""
    q = _q(Product, db, hub_id).filter(Product.is_active == True)  # noqa: E712

    if search:
        q = q.filter(or_(
            Product.name.ilike(f"%{search}%"),
            Product.sku.ilike(f"%{search}%"),
        ))

    # Sort
    sort_map = {
        "name": Product.name, "-name": Product.name.desc(),
        "sku": Product.sku, "-sku": Product.sku.desc(),
        "price": Product.price, "-price": Product.price.desc(),
        "stock": Product.stock, "-stock": Product.stock.desc(),
        "id": Product.id, "-id": Product.created_at.desc(),
    }
    q = q.order_by(sort_map.get(order_by, Product.created_at.desc()))

    total_count = await q.count()
    products = await q.options(
        selectinload(Product.categories),
    ).offset((page - 1) * per_page).limit(per_page).all()

    has_next = (page * per_page) < total_count

    config = await _get_config(db, hub_id)

    context = {
        "products": products,
        "has_next": has_next,
        "next_page": page + 1 if has_next else None,
        "is_first_page": page == 1,
        "total_count": total_count,
        "start_index": (page - 1) * per_page + 1 if total_count > 0 else 0,
        "end_index": min(page * per_page, total_count),
        "page_number": page,
        "barcode_enabled": config.barcode_enabled,
        "search": search,
        "order_by": order_by,
        "per_page": per_page,
    }

    hx_target = request.headers.get("HX-Target", "")

    if page > 1:
        return context  # @htmx_view picks products_rows_infinite partial

    if hx_target == "products-table-container":
        return context  # @htmx_view picks products_table_partial

    return context


# ============================================================================
# Product AJAX list (JSON)
# ============================================================================

@router.get("/products/api/list")
async def product_list_ajax(
    request: Request, db: DbSession, hub_id: HubId,
    search: str = "", page: int = 1, per_page: int = 10,
):
    """Return products as JSON for dynamic table."""
    q = _q(Product, db, hub_id).filter(Product.is_active == True)  # noqa: E712

    if search:
        q = q.filter(or_(
            Product.name.ilike(f"%{search}%"),
            Product.sku.ilike(f"%{search}%"),
        ))

    total = await q.count()
    products = await q.options(
        selectinload(Product.categories),
    ).offset((page - 1) * per_page).limit(per_page).all()

    products_data = []
    for product in products:
        cats = [{"id": str(c.id), "name": c.name} for c in product.categories]
        products_data.append({
            "id": str(product.id),
            "name": product.name,
            "sku": product.sku,
            "categories": cats,
            "category": cats[0]["name"] if cats else "Sin categoria",
            "category_ids": [c["id"] for c in cats],
            "price": float(product.price),
            "cost": float(product.cost),
            "stock": product.stock,
            "low_stock_threshold": product.low_stock_threshold,
            "is_low_stock": product.is_low_stock,
            "image": product.get_image_path(),
            "initial": product.get_initial(),
            "profit_margin": product.profit_margin,
        })

    return JSONResponse({
        "products": products_data,
        "total": total,
        "pages": (total + per_page - 1) // per_page,
        "current_page": page,
    })


# ============================================================================
# Product Create
# ============================================================================

@router.get("/products/create")
@htmx_view(module_id="inventory", view_id="product_create", permissions="inventory.add_product")
async def product_create_form(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Show product create form."""
    categories = await _q(Category, db, hub_id).filter(
        Category.is_active == True  # noqa: E712
    ).order_by(Category.order, Category.name).all()

    return {
        "categories": categories,
        "mode": "create",
        "readonly": False,
    }


@router.post("/products/create")
async def product_create_post(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Create a new product."""
    form = await request.form()

    try:
        price_str = form.get("price", "").strip()
        if not price_str:
            raise ValueError("Price is required")

        cost_str = form.get("cost", "0").strip()
        stock_str = form.get("stock", "0").strip()
        threshold_str = form.get("low_stock_threshold", "10").strip()
        product_type = form.get("product_type", "physical")
        tax_class_id_str = form.get("tax_class_id", "").strip()

        async with atomic(db) as session:
            product = Product(
                hub_id=hub_id,
                name=form.get("name", ""),
                sku=form.get("sku", ""),
                ean13=form.get("ean13", "").strip() or None,
                description=form.get("description", ""),
                product_type=product_type,
                price=Decimal(price_str),
                cost=Decimal(cost_str) if cost_str else Decimal("0"),
                stock=int(stock_str) if stock_str else 0,
                low_stock_threshold=int(threshold_str) if threshold_str else 10,
                tax_class_id=uuid.UUID(tax_class_id_str) if tax_class_id_str else None,
            )

            # Handle image upload
            image_file = form.get("image")
            if image_file and hasattr(image_file, "read"):
                # TODO: save via media service
                pass

            session.add(product)
            await session.flush()

            # Handle categories
            category_names = form.getlist("category_names[]") or form.getlist("category_names")
            if len(category_names) == 1 and "," in category_names[0]:
                category_names = category_names[0].split(",")
            category_names = [n.strip().capitalize() for n in category_names if n.strip()]
            if category_names:
                cats = await _q(Category, session, hub_id).filter(
                    Category.name.in_(category_names)
                ).all()
                product.categories = cats

        return htmx_redirect("/m/inventory/products")

    except Exception as e:
        categories = await _q(Category, db, hub_id).filter(
            Category.is_active == True  # noqa: E712
        ).order_by(Category.order, Category.name).all()
        return {
            "categories": categories,
            "mode": "create",
            "readonly": False,
            "error_message": str(e),
        }


# ============================================================================
# Product View (read-only)
# ============================================================================

@router.get("/products/{pk}")
@htmx_view(module_id="inventory", view_id="product_view", permissions="inventory.view_product")
async def product_view(
    request: Request, pk: uuid.UUID, db: DbSession, user: CurrentUser, hub_id: HubId,
):
    """View product in read-only mode."""
    product = await _q(Product, db, hub_id).options(
        selectinload(Product.categories),
    ).get(pk)
    if product is None:
        return JSONResponse({"error": "Product not found"}, status_code=404)

    categories = await _q(Category, db, hub_id).filter(
        Category.is_active == True  # noqa: E712
    ).order_by(Category.order, Category.name).all()

    return {
        "product": product,
        "categories": categories,
        "mode": "view",
        "readonly": True,
    }


# ============================================================================
# Product Edit
# ============================================================================

@router.get("/products/{pk}/edit")
@htmx_view(module_id="inventory", view_id="product_edit", permissions="inventory.change_product")
async def product_edit_form(
    request: Request, pk: uuid.UUID, db: DbSession, user: CurrentUser, hub_id: HubId,
):
    """Show product edit form."""
    product = await _q(Product, db, hub_id).options(
        selectinload(Product.categories),
    ).get(pk)
    if product is None:
        return JSONResponse({"error": "Product not found"}, status_code=404)

    categories = await _q(Category, db, hub_id).filter(
        Category.is_active == True  # noqa: E712
    ).order_by(Category.order, Category.name).all()

    return {
        "product": product,
        "categories": categories,
        "mode": "edit",
        "readonly": False,
    }


@router.post("/products/{pk}/edit")
async def product_edit_post(
    request: Request, pk: uuid.UUID, db: DbSession, user: CurrentUser, hub_id: HubId,
):
    """Update an existing product."""
    product = await _q(Product, db, hub_id).options(
        selectinload(Product.categories),
    ).get(pk)
    if product is None:
        return JSONResponse({"error": "Product not found"}, status_code=404)

    form = await request.form()

    try:
        price_str = form.get("price", "").strip()
        if not price_str:
            raise ValueError("Price is required")

        cost_str = form.get("cost", "0").strip()
        stock_str = form.get("stock", "0").strip()
        threshold_str = form.get("low_stock_threshold", "10").strip()

        async with atomic(db) as session:
            product.name = form.get("name", product.name)
            product.sku = form.get("sku", product.sku)
            product.ean13 = form.get("ean13", "").strip() or None
            product.description = form.get("description", "")
            product.product_type = form.get("product_type", "physical")
            product.price = Decimal(price_str)
            product.cost = Decimal(cost_str) if cost_str else Decimal("0")
            product.stock = int(stock_str) if stock_str else 0
            product.low_stock_threshold = int(threshold_str) if threshold_str else 10

            tax_class_id_str = form.get("tax_class_id", "").strip()
            product.tax_class_id = uuid.UUID(tax_class_id_str) if tax_class_id_str else None

            # Handle image upload
            image_file = form.get("image")
            if image_file and hasattr(image_file, "read"):
                # TODO: save via media service
                pass

            # Update categories
            category_names = form.getlist("category_names[]") or form.getlist("category_names")
            if len(category_names) == 1 and "," in category_names[0]:
                category_names = category_names[0].split(",")
            category_names = [n.strip().capitalize() for n in category_names if n.strip()]
            if category_names:
                cats = await _q(Category, session, hub_id).filter(
                    Category.name.in_(category_names)
                ).all()
                product.categories = cats
            else:
                product.categories = []

        return htmx_redirect("/m/inventory/products")

    except Exception as e:
        categories = await _q(Category, db, hub_id).filter(
            Category.is_active == True  # noqa: E712
        ).order_by(Category.order, Category.name).all()
        return {
            "product": product,
            "categories": categories,
            "mode": "edit",
            "readonly": False,
            "error_message": str(e),
        }


# ============================================================================
# Product Delete
# ============================================================================

@router.post("/products/{pk}/delete")
@router.delete("/products/{pk}/delete")
async def product_delete(
    request: Request, pk: uuid.UUID, db: DbSession, user: CurrentUser, hub_id: HubId,
):
    """Delete a product."""
    try:
        await _q(Product, db, hub_id).delete(pk)
        return JSONResponse({"success": True, "message": "Product deleted successfully"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=400)


# ============================================================================
# Product Barcode
# ============================================================================

@router.get("/products/{product_id}/barcode")
async def generate_barcode(
    request: Request, product_id: uuid.UUID, db: DbSession, user: CurrentUser, hub_id: HubId,
    type: str = "sku",
):
    """Generate barcode SVG for a product."""
    from .barcode_utils import generate_barcode_svg

    config = await _get_config(db, hub_id)
    if not config.barcode_enabled:
        return JSONResponse({"success": False, "error": "Barcode generation is disabled"}, status_code=403)

    product = await _q(Product, db, hub_id).get(product_id)
    if product is None:
        return JSONResponse({"success": False, "error": "Product not found"}, status_code=404)

    try:
        if type == "ean13":
            if not product.ean13:
                return Response(
                    content='<svg><text x="50%" y="50%" text-anchor="middle" fill="red">No EAN-13 available</text></svg>',
                    media_type="image/svg+xml",
                )
            svg_content = generate_barcode_svg(product.ean13, format_type="ean13")
        else:
            svg_content = generate_barcode_svg(product.sku, format_type="code128")

        return Response(content=svg_content, media_type="image/svg+xml")

    except ValueError as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)
    except Exception:
        return JSONResponse({"success": False, "error": "Error generating barcode"}, status_code=500)


# ============================================================================
# Product Export CSV
# ============================================================================

@router.get("/products/export/csv")
async def export_csv(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Export products to CSV."""
    products = await _q(Product, db, hub_id).filter(
        Product.is_active == True  # noqa: E712
    ).options(selectinload(Product.categories)).all()

    output = io.StringIO()
    output.write("\ufeff")  # BOM
    writer = csv_mod.writer(output)
    writer.writerow(["sku", "name", "description", "categories", "price", "cost", "stock", "low_stock_threshold", "ean13"])

    for product in products:
        cats = ", ".join(c.name for c in product.categories) if product.categories else ""
        writer.writerow([
            product.sku, product.name, product.description, cats,
            float(product.price), float(product.cost), product.stock,
            product.low_stock_threshold, product.ean13 or "",
        ])

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="productos.csv"'},
    )


# ============================================================================
# Product Import CSV
# ============================================================================

@router.post("/products/import/csv")
async def import_csv(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Import products from CSV."""
    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse({"success": False, "message": "No file provided"}, status_code=400)

    content = await file.read()
    text = content.decode("utf-8-sig")
    reader = csv_mod.DictReader(io.StringIO(text))

    created = 0
    updated = 0
    errors = []

    async with atomic(db) as session:
        for row in reader:
            try:
                sku = row.get("sku", "").strip()
                if not sku:
                    continue

                name = row.get("name", "").strip()
                if not name:
                    errors.append(f"Row {reader.line_num}: Name is required")
                    continue

                price = Decimal(row.get("price", "0"))
                cost = Decimal(row.get("cost", "0"))
                stock = int(row.get("stock", "0"))

                if price < 0 or cost < 0 or stock < 0:
                    errors.append(f"Row {reader.line_num}: Negative values not allowed")
                    continue

                # Get or create categories
                categories_str = row.get("categories", "").strip()
                cats = []
                if categories_str:
                    cat_names = [n.strip() for n in categories_str.split(",") if n.strip()]
                    for cn in cat_names:
                        existing = await _q(Category, session, hub_id).filter(
                            Category.name == cn
                        ).first()
                        if not existing:
                            existing = Category(hub_id=hub_id, name=cn, icon="cube-outline")
                            session.add(existing)
                            await session.flush()
                        cats.append(existing)

                # Find existing by SKU
                existing_product = await _q(Product, session, hub_id).filter(
                    Product.sku == sku
                ).first()

                if existing_product:
                    existing_product.name = name
                    existing_product.description = row.get("description", "")
                    existing_product.price = price
                    existing_product.cost = cost
                    existing_product.stock = stock
                    existing_product.low_stock_threshold = int(row.get("low_stock_threshold", "10"))
                    existing_product.ean13 = row.get("ean13", "") or None
                    if cats:
                        existing_product.categories = cats
                    updated += 1
                else:
                    product = Product(
                        hub_id=hub_id,
                        sku=sku, name=name,
                        description=row.get("description", ""),
                        price=price, cost=cost, stock=stock,
                        low_stock_threshold=int(row.get("low_stock_threshold", "10")),
                        ean13=row.get("ean13", "") or None,
                    )
                    session.add(product)
                    await session.flush()
                    if cats:
                        product.categories = cats
                    created += 1

            except Exception as e:
                errors.append(f"Row {reader.line_num}: {e!s}")

    return JSONResponse({
        "success": True,
        "message": f"{created} products created, {updated} updated",
        "created": created, "updated": updated, "errors": errors,
    })


# ============================================================================
# Product Import Excel
# ============================================================================

@router.post("/products/import/excel")
async def import_excel(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Import products from Excel file."""
    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse({"error": "No file provided"}, status_code=400)

    try:
        import openpyxl

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        header_map = {h.lower(): i for i, h in enumerate(headers) if h}

        created_count = 0
        updated_count = 0
        errors = []

        async with atomic(db) as session:
            def _get_value(row, header_name, default=""):
                idx = header_map.get(header_name.lower())
                if idx is not None and idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val is not None else default
                return default

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    name = _get_value(row, "Name")
                    sku = _get_value(row, "SKU")
                    if not name or not sku:
                        errors.append(f"Row {row_num}: Name and SKU are required")
                        continue

                    price = Decimal(_get_value(row, "Price", "0"))
                    cost = Decimal(_get_value(row, "Cost", "0"))
                    stock = int(float(_get_value(row, "Stock", "0")))
                    low_stock_threshold = int(float(_get_value(row, "Low Stock Threshold", "10")))
                    ean13 = _get_value(row, "EAN-13")

                    existing = await _q(Product, session, hub_id).filter(
                        Product.sku == sku
                    ).first()

                    if existing:
                        existing.name = name
                        existing.description = _get_value(row, "Description")
                        existing.price = price
                        existing.cost = cost
                        existing.stock = stock
                        existing.low_stock_threshold = low_stock_threshold
                        if ean13:
                            existing.ean13 = ean13
                        updated_count += 1
                        product = existing
                    else:
                        product = Product(
                            hub_id=hub_id, name=name, sku=sku,
                            description=_get_value(row, "Description"),
                            price=price, cost=cost, stock=stock,
                            low_stock_threshold=low_stock_threshold,
                            ean13=ean13 if ean13 else None,
                        )
                        session.add(product)
                        await session.flush()
                        created_count += 1

                    # Handle categories
                    categories_str = _get_value(row, "Categories")
                    if categories_str:
                        product.categories = []
                        cat_names = [c.strip().capitalize() for c in categories_str.split(",") if c.strip()]
                        for cn in cat_names:
                            cat = await _q(Category, session, hub_id).filter(
                                func.lower(Category.name) == cn.lower()
                            ).first()
                            if not cat:
                                cat = Category(hub_id=hub_id, name=cn)
                                session.add(cat)
                                await session.flush()
                            product.categories.append(cat)

                except Exception as e:
                    errors.append(f"Row {row_num}: {e!s}")

        message = f"Import completed: {created_count} products created, {updated_count} updated"
        if errors:
            message += f". {len(errors)} errors occurred."
        return JSONResponse({
            "success": True, "message": message,
            "created": created_count, "updated": updated_count, "errors": errors,
        })

    except ImportError:
        return JSONResponse({"error": "openpyxl not installed"}, status_code=400)
    except Exception as e:
        return JSONResponse({"error": f"Error processing file: {e!s}"}, status_code=500)


# ============================================================================
# Categories
# ============================================================================

@router.get("/categories")
@htmx_view(module_id="inventory", view_id="categories", permissions="inventory.view_category")
async def categories_index(
    request: Request, db: DbSession, user: CurrentUser, hub_id: HubId,
    search: str = "", order_by: str = "order", page: int = 1, per_page: int = 25,
):
    """Categories list with infinite scroll."""
    q = _q(Category, db, hub_id).filter(Category.is_active == True)  # noqa: E712

    if search:
        q = q.filter(or_(
            Category.name.ilike(f"%{search}%"),
            Category.description.ilike(f"%{search}%"),
        ))

    q = q.options(selectinload(Category.products))

    sort_map = {
        "name": Category.name, "-name": Category.name.desc(),
        "order": Category.order, "-order": Category.order.desc(),
    }
    q = q.order_by(sort_map.get(order_by, Category.order), Category.name)

    total_count = await q.count()
    categories = await q.offset((page - 1) * per_page).limit(per_page).all()
    has_next = (page * per_page) < total_count

    context = {
        "categories": categories,
        "has_next": has_next,
        "next_page": page + 1 if has_next else None,
        "is_first_page": page == 1,
        "total_count": total_count,
        "search": search,
        "order_by": order_by,
        "per_page": per_page,
    }

    hx_target = request.headers.get("HX-Target", "")
    page_num = page

    if page_num > 1:
        return context

    if hx_target == "categories-table-container":
        return context

    return context


@router.get("/categories/api")
async def categories_list_api(
    request: Request, db: DbSession, hub_id: HubId,
    search: str = "", id: str = "", page: int = 1, per_page: int = 10,
):
    """Categories as JSON."""
    q = _q(Category, db, hub_id).filter(Category.is_active == True)  # noqa: E712

    if id:
        q = q.filter(Category.id == uuid.UUID(id))

    if search:
        q = q.filter(or_(
            Category.name.ilike(f"%{search}%"),
            Category.description.ilike(f"%{search}%"),
        ))

    q = q.options(selectinload(Category.products))
    total = await q.count()
    categories = await q.order_by(Category.order, Category.name).offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    return JSONResponse({
        "success": True,
        "categories": [
            {
                "id": str(c.id), "name": c.name, "icon": c.icon,
                "color": c.color, "description": c.description,
                "order": c.order, "image": c.get_image_url(),
                "initial": c.get_initial(),
                "product_count": c.product_count,
            }
            for c in categories
        ],
        "total": total,
    })


@router.get("/categories/create")
@htmx_view(module_id="inventory", view_id="category_create", permissions="inventory.add_category")
async def category_create_form(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Show category create form."""
    return {"mode": "create"}


@router.post("/categories/create")
async def category_create_post(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Create a new category."""
    form = await request.form()
    try:
        tax_class_id_str = form.get("tax_class_id", "").strip()

        async with atomic(db) as session:
            category = Category(
                hub_id=hub_id,
                name=form.get("name", ""),
                description=form.get("description", ""),
                icon=form.get("icon", "cube-outline"),
                color=form.get("color", "#3880ff"),
                order=int(form.get("order", 0)),
                tax_class_id=uuid.UUID(tax_class_id_str) if tax_class_id_str else None,
            )
            session.add(category)

        return JSONResponse({"success": True, "message": "Category created", "category_id": str(category.id)})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=400)


@router.get("/categories/edit/{pk}")
@htmx_view(module_id="inventory", view_id="category_edit", permissions="inventory.change_category")
async def category_edit_form(
    request: Request, pk: uuid.UUID, db: DbSession, user: CurrentUser, hub_id: HubId,
):
    """Show category edit form."""
    category = await _q(Category, db, hub_id).get(pk)
    if category is None:
        return JSONResponse({"error": "Category not found"}, status_code=404)
    return {"category": category, "mode": "edit"}


@router.post("/categories/edit/{pk}")
async def category_edit_post(
    request: Request, pk: uuid.UUID, db: DbSession, user: CurrentUser, hub_id: HubId,
):
    """Update a category."""
    category = await _q(Category, db, hub_id).get(pk)
    if category is None:
        return JSONResponse({"error": "Category not found"}, status_code=404)

    form = await request.form()
    try:
        category.name = form.get("name", category.name)
        category.description = form.get("description", "")
        category.icon = form.get("icon", "cube-outline")
        category.color = form.get("color", "#3880ff")
        category.order = int(form.get("order", 0))

        tax_class_id_str = form.get("tax_class_id", "").strip()
        category.tax_class_id = uuid.UUID(tax_class_id_str) if tax_class_id_str else None

        await db.flush()

        return JSONResponse({"success": True, "message": "Category updated"})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=400)


@router.post("/categories/delete/{pk}")
async def category_delete(
    request: Request, pk: uuid.UUID, db: DbSession, user: CurrentUser, hub_id: HubId,
):
    """Delete a category."""
    category = await _q(Category, db, hub_id).options(
        selectinload(Category.products),
    ).get(pk)
    if category is None:
        return JSONResponse({"error": "Category not found"}, status_code=404)

    active_products = len([p for p in category.products if p.is_active])
    if active_products > 0:
        return JSONResponse({
            "success": False,
            "message": f"Cannot delete category with {active_products} active products",
        }, status_code=400)

    await _q(Category, db, hub_id).delete(pk)
    return JSONResponse({"success": True, "message": "Category deleted"})


# ============================================================================
# Category Export
# ============================================================================

@router.get("/categories/export/csv")
async def export_categories_csv(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Export categories to CSV."""
    categories = await _q(Category, db, hub_id).filter(
        Category.is_active == True  # noqa: E712
    ).options(selectinload(Category.products)).order_by(Category.order, Category.name).all()

    output = io.StringIO()
    writer = csv_mod.writer(output)
    writer.writerow(["ID", "Name", "Description", "Icon", "Color", "Order", "Product Count"])

    for cat in categories:
        writer.writerow([
            str(cat.id), cat.name, cat.description or "", cat.icon,
            cat.color, cat.order, cat.product_count,
        ])

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="categories.csv"'},
    )


@router.get("/categories/export/excel")
async def export_categories_excel(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Export categories to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    categories = await _q(Category, db, hub_id).filter(
        Category.is_active == True  # noqa: E712
    ).options(selectinload(Category.products)).order_by(Category.order, Category.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Categories"

    headers = ["ID", "Name", "Description", "Icon", "Color", "Order", "Product Count"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for cat in categories:
        ws.append([str(cat.id), cat.name, cat.description or "", cat.icon, cat.color, cat.order, cat.product_count])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="categories.xlsx"'},
    )


# ============================================================================
# Category Import
# ============================================================================

@router.post("/categories/import/csv")
async def import_categories_csv(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Import categories from CSV."""
    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse({"error": "No file provided"}, status_code=400)

    content = await file.read()
    text = content.decode("utf-8-sig")
    reader = csv_mod.DictReader(io.StringIO(text))

    created_count = 0
    updated_count = 0
    errors = []

    async with atomic(db) as session:
        for row_num, row in enumerate(reader, start=2):
            try:
                name = row.get("Name", "").strip()
                if not name:
                    errors.append(f"Row {row_num}: Name is required")
                    continue

                normalized_name = name.capitalize()
                existing = await _q(Category, session, hub_id).filter(
                    func.lower(Category.name) == normalized_name.lower()
                ).first()

                if existing:
                    existing.description = row.get("Description", "").strip()
                    existing.icon = row.get("Icon", "pricetag-outline").strip()
                    existing.color = row.get("Color", "primary").strip()
                    existing.order = int(row.get("Order", 100))
                    updated_count += 1
                else:
                    cat = Category(
                        hub_id=hub_id,
                        name=normalized_name,
                        description=row.get("Description", "").strip(),
                        icon=row.get("Icon", "pricetag-outline").strip(),
                        color=row.get("Color", "primary").strip(),
                        order=int(row.get("Order", 100)),
                    )
                    session.add(cat)
                    created_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {e!s}")

    message = f"Import completed: {created_count} categories created, {updated_count} updated"
    if errors:
        message += f". {len(errors)} errors."
    return JSONResponse({
        "success": True, "message": message,
        "created": created_count, "updated": updated_count, "errors": errors,
    })


@router.post("/categories/import/excel")
async def import_categories_excel(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Import categories from Excel."""
    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse({"error": "No file provided"}, status_code=400)

    try:
        import openpyxl

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        header_map = {h.lower(): i for i, h in enumerate(headers) if h}

        created_count = 0
        updated_count = 0
        errors = []

        def _get_cat_value(row, header_name, default=""):
            idx = header_map.get(header_name.lower())
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else default
            return default

        async with atomic(db) as session:
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    name = _get_cat_value(row, "Name")
                    if not name:
                        errors.append(f"Row {row_num}: Name is required")
                        continue

                    normalized_name = name.capitalize()
                    existing = await _q(Category, session, hub_id).filter(
                        func.lower(Category.name) == normalized_name.lower()
                    ).first()

                    if existing:
                        existing.description = _get_value(row, "Description")
                        existing.icon = _get_value(row, "Icon", "pricetag-outline")
                        existing.color = _get_value(row, "Color", "primary")
                        existing.order = int(float(_get_value(row, "Order", "100")))
                        updated_count += 1
                    else:
                        cat = Category(
                            hub_id=hub_id,
                            name=normalized_name,
                            description=_get_value(row, "Description"),
                            icon=_get_value(row, "Icon", "pricetag-outline"),
                            color=_get_value(row, "Color", "primary"),
                            order=int(float(_get_value(row, "Order", "100"))),
                        )
                        session.add(cat)
                        created_count += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {e!s}")

        message = f"Import completed: {created_count} categories created, {updated_count} updated"
        if errors:
            message += f". {len(errors)} errors."
        return JSONResponse({
            "success": True, "message": message,
            "created": created_count, "updated": updated_count, "errors": errors,
        })

    except ImportError:
        return JSONResponse({"error": "openpyxl not installed"}, status_code=400)
    except Exception as e:
        return JSONResponse({"error": f"Error processing file: {e!s}"}, status_code=500)


# ============================================================================
# Reports
# ============================================================================

@router.get("/reports")
@htmx_view(module_id="inventory", view_id="reports")
async def reports_view(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Inventory reports and statistics."""
    q = _q(Product, db, hub_id)
    active_q = q.filter(Product.is_active == True)  # noqa: E712

    total_products = await active_q.count()
    products_in_stock = await active_q.filter(Product.stock > 0).count()
    products_out_of_stock = await active_q.filter(Product.stock == 0).count()
    products_low_stock = await active_q.filter(
        Product.stock <= Product.low_stock_threshold, Product.stock > 0
    ).count()

    total_inventory_value = await active_q.sum(Product.stock * Product.price) or 0
    total_cost_value = await active_q.sum(Product.stock * Product.cost) or 0
    total_units = await active_q.sum(Product.stock) or 0
    total_categories = await _q(Category, db, hub_id).filter(
        Category.is_active == True  # noqa: E712
    ).count()

    # Category stats
    categories = await _q(Category, db, hub_id).filter(
        Category.is_active == True  # noqa: E712
    ).options(selectinload(Category.products)).order_by(Category.order, Category.name).all()

    category_stats = []
    for cat in categories:
        active_products = [p for p in cat.products if p.is_active]
        if active_products:
            total_stock = sum(p.stock for p in active_products)
            total_value = sum(p.stock * p.price for p in active_products)
            category_stats.append({
                "name": cat.name, "icon": cat.icon, "color": cat.color,
                "product_count": len(active_products),
                "total_stock": total_stock, "total_value": total_value,
            })

    # Top products by stock value
    top_by_value = await active_q.filter(Product.stock > 0).order_by(
        (Product.stock * Product.price).desc()
    ).limit(10).all()
    for p in top_by_value:
        p.stock_value = p.stock * p.price

    # Top products by stock quantity
    top_by_stock = await active_q.filter(Product.stock > 0).order_by(
        Product.stock.desc()
    ).limit(10).all()

    # Critical stock
    critical = await active_q.filter(
        Product.stock <= Product.low_stock_threshold, Product.stock > 0
    ).order_by(Product.stock).limit(20).all()

    return {
        "total_products": total_products,
        "products_in_stock": products_in_stock,
        "products_out_of_stock": products_out_of_stock,
        "products_low_stock": products_low_stock,
        "total_inventory_value": total_inventory_value,
        "total_cost_value": total_cost_value,
        "total_units": total_units,
        "total_categories": total_categories,
        "category_stats": category_stats,
        "top_products_by_value": top_by_value,
        "top_products_by_stock": top_by_stock,
        "critical_stock_products": critical,
    }


# ============================================================================
# Settings
# ============================================================================

@router.get("/settings")
@htmx_view(module_id="inventory", view_id="settings", permissions="inventory.manage_settings")
async def settings_view(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Settings page."""
    config = await _get_config(db, hub_id)
    return {"config": config}


@router.post("/settings")
async def settings_save(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Save settings (JSON body)."""
    import json as json_module

    config = await _get_config(db, hub_id)
    body = await request.body()
    data = json_module.loads(body)

    config.allow_negative_stock = data.get("allow_negative_stock", False)
    config.low_stock_alert_enabled = data.get("low_stock_alert_enabled", True)
    config.barcode_enabled = data.get("barcode_enabled", True)
    await db.flush()

    return JSONResponse({"success": True, "message": "Settings saved"})


@router.post("/settings/toggle")
async def settings_toggle(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Toggle individual setting."""
    form = await request.form()
    name = form.get("name")
    value = form.get("value", "").lower()

    if not name:
        return Response(status_code=400)

    config = await _get_config(db, hub_id)

    if not hasattr(config, name):
        return Response(status_code=400)

    setattr(config, name, value == "true")
    await db.flush()

    return Response(
        status_code=204,
        headers={"HX-Trigger": '{"showToast": {"message": "Setting saved", "color": "success"}}'},
    )


@router.post("/settings/reset")
async def settings_reset(request: Request, db: DbSession, user: CurrentUser, hub_id: HubId):
    """Reset settings to defaults."""
    config = await _get_config(db, hub_id)
    config.allow_negative_stock = False
    config.low_stock_alert_enabled = True
    config.barcode_enabled = True
    await db.flush()

    return {"config": config}
